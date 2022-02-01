from tokenize import Name
from cv2 import waitKey
from flask import Flask, render_template, url_for, request, redirect, session, jsonify
from flask_sqlalchemy import SQLAlchemy
import json
import os
import pandas as pd
import openpyxl 
from openpyxl_image_loader import SheetImageLoader
from flask_mysqldb import MySQL
import numpy as np
from PIL import Image as im
import cv2
import base64

import mysql.connector

with open('config.json', 'r') as c:
    params = json.load(c) ["params"]

local_server = True
app = Flask(__name__)
if(local_server):
    app.config['SQLALCHEMY_DATABASE_URI'] = params['local_uri']
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = params['prod_uri']


db = SQLAlchemy(app)

class Register(db.Model):
    # sno,Name,Gender,Type,email,password
    sno = db.Column(db.Integer, primary_key=True)
    Name = db.Column(db.String(80), nullable=False)
    Gender = db.Column(db.String(80), nullable=False)
    Type = db.Column(db.String(80), nullable=False)
    email = db.Column(db.String(80), nullable=False)
    password = db.Column(db.String(80), nullable=False)


app.secret_key=os.urandom(24)

app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'register_db'
mysql = MySQL(app)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/register')
def register_page():
    return render_template('register.html')

@app.route('/rawattech')
def home():
    if 'user_id' in session:
        return render_template('home.html')
    else:
        return redirect('/')

@app.route('/logout')
def logout():
    session.pop('user_id')
    return redirect('/')

@app.route('/login_validation', methods=['POST'])
def login_validation():
    email = request.form.get('email')
    password = request.form.get('password')
    cursor = mysql.connection.cursor()
    cursor.execute('SELECT * FROM register WHERE Name = %s AND password = %s', (email, password,))
    register = cursor.fetchall()
    if len(register)>0:
        session['user_id']=register[0][0]
        session['user_name']=email
        return redirect('/rawattech')
    else:
        return redirect('/')
    mysql.connection.commit()
    cursor.close()

@app.route('/add_user',methods=['GET' , 'POST'])
def add_user():
    if(request.method=='POST'):
        name=request.form.get('your_name')
        gender=request.form.get('inlineRadioOptions')
        type=request.form.get('type_of_account')
        email=request.form.get('email')
        password=request.form.get('password')

        entry = Register(Name=name, Gender=gender, Type=type, email=email, password=password )
        db.session.add(entry)
        db.session.commit()
        
    return redirect('/')   
    # name=request.form.get('your_name')
    # gender=request.form.get('inlineRadioOptions')
    # type=request.form.get('type_of_account')
    # email=request.form.get('email')
    # password=request.form.get('password')

    # return redirect('/')
    # mysql.connection.commit()
    # cursor.close()
    

@app.route("/data", methods=['POST', 'GET'])
def data():
        # file = request.form.get('Daily Tasks.xlsx')
        wb = openpyxl.load_workbook('Daily Tasks.xlsx')
        ws = wb['Jan_2022']
        headers =[]
        for cell in ws[1]:
            headers.append(cell.value)
        sheet_data = []
        row_count = ws.max_row
        for i in range(2,7):
            for cell in ws[i]:
                sheet_data.append(cell.value)
        sheet_data = tuple(sheet_data)
        res = tuple(sheet_data[x:x + 4] 
            for x in range(0, len(sheet_data), 4))
        data = ws['A2'].value
        # data = pd.read_excel(file)
        return render_template('data.html', headers=headers, sheet_data=res)
@app.route("/upload", methods=['GET', 'POST'])
def upload_file():
    # if request.method == 'POST':
    #     return jsonify({"result": request.get_array(field_name='file')})
    #loading the Excel File and the sheet
    # file1 = request.form['file']
    pxl_doc = openpyxl.load_workbook('japanese.xlsx')
    sheet = pxl_doc['Sheet1']

    #calling the image_loader
    image_loader = SheetImageLoader(sheet)

    #get the image (put the cell you need instead of 'A1')
    image = image_loader.get('A3')
    
    img_arr = np.array(image)
    img_arr1 = img_arr.tobytes()
    output = img_arr1.decode('utf-16')
    msg = base64.b64decode(output)
    # output = img_arr1.decode()
    # msg = base64.b64decode(output)
    # img_arr = img_arr.reshape(img_arr.shape[0], img_arr.shape[1], 3)
    # img_arr = cv2.cvtColor(img_arr, cv2.COLOR_BGR2RGB)
    # encoded_image = cv2.imencode('.png', img_arr)
    # content2 = encoded_image.tobytes()

    #showing the image
    return render_template('data.html', data=msg)
    
if __name__ == "__main__":
    app.run(debug=True)
# from . import routes  
